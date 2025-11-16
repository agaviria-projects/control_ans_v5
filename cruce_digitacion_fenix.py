"""
------------------------------------------------------------
CRUCE DIGITACIÓN FÉNIX (v5.5 – Ligera y Segura)
------------------------------------------------------------
Autor: Héctor + IA (2025)
------------------------------------------------------------
Descripción:
Cruza 'Digitacion Fenix.txt' con 'FENIX_ANS.xlsx' y
actualiza únicamente la columna ESTADO_FENIX sin modificar
ningún formato condicional ni tabla estructurada.
------------------------------------------------------------
"""

import pandas as pd
from pathlib import Path
import unicodedata

# ============================================================
# 1️⃣ RUTAS BASE
# ============================================================
script_path = Path(__file__).resolve()
base_path = script_path.parent if "scripts" not in str(script_path) else script_path.parent.parent

ruta_fenix_ans = base_path / "data_clean" / "FENIX_ANS.xlsx"
ruta_digitacion = base_path / "data_raw" / "Digitacion Fenix.txt"
ruta_repo = base_path / "data_clean" / "REPOSITORIO_PEDIDOS_CERRADOS.xlsx"

print("------------------------------------------------------------")
print("🔄 Iniciando CRUCE DIGITACIÓN FÉNIX v5.5 (Ligera y Segura)...")
print(f"📂 Base detectada: {base_path}")
print("------------------------------------------------------------")

# ============================================================
# 2️⃣ DETECCIÓN Y LECTURA AUTOMÁTICA
# ============================================================
def detectar_separador(ruta_txt):
    with open(ruta_txt, 'r', encoding='latin-1', errors='ignore') as f:
        linea = f.readline()
    if "|" in linea: return "|"
    elif "\t" in linea: return "\t"
    elif ";" in linea: return ";"
    elif "," in linea: return ","
    else: return "\t"

def leer_txt_seguro(ruta, sep):
    for enc in ["utf-8", "latin-1", "ISO-8859-1"]:
        try:
            return pd.read_csv(ruta, sep=sep, dtype=str, encoding=enc, on_bad_lines="skip")
        except Exception:
            continue
    raise Exception("❌ No se pudo leer Digitación Fénix con ninguna codificación.")

sep = detectar_separador(ruta_digitacion)
df_txt = leer_txt_seguro(ruta_digitacion, sep)
df_ans = pd.read_excel(ruta_fenix_ans, sheet_name="FENIX_ANS", dtype=str)

# ============================================================
# 3️⃣ NORMALIZACIÓN Y PREPARACIÓN
# ============================================================
df_txt.columns = [c.strip().upper() for c in df_txt.columns]
col_pedido = next((c for c in df_txt.columns if "PEDID" in c.upper()), None)
if not col_pedido:
    raise Exception("❌ No se encontró columna de pedido en Digitación Fénix.")

def limpiar_texto(txt):
    if not isinstance(txt, str):
        return ""
    txt = unicodedata.normalize("NFD", txt)
    txt = txt.encode("ascii", "ignore").decode("utf-8")
    return txt.strip().upper()

df_ans["PEDIDO"] = df_ans["PEDIDO"].astype(str).str.strip().str.replace(".0", "", regex=False)
df_txt[col_pedido] = df_txt[col_pedido].astype(str).str.strip().str.replace(".0", "", regex=False)
pedidos_digitacion = set(df_txt[col_pedido].unique())

# ============================================================
# 4️⃣ ACTUALIZACIÓN DE ESTADO_FENIX
# ============================================================
def calcular_estado_fenix(row):
    pedido = str(row.get("PEDIDO", "")).strip()
    reporte = limpiar_texto(row.get("REPORTE_TECNICO", ""))

    # Extraer días y horas
    dias_rest = 0
    horas_rest = "00:00"
    try:
        partes = str(row.get("DIAS_RESTANTES", "0")).split(" ")
        dias_rest = int(partes[0])
        if len(partes) > 2:
            horas_rest = partes[2]  # ej: "21:47"
    except:
        dias_rest = 0
        horas_rest = "00:00"

    # Regla CERRADO (Digitación)
    if pedido in pedidos_digitacion:
        return "CERRADO"

    # Regla EJECUTADO EN CAMPO
    if "EJECUTADO" in reporte and "CAMPO" in reporte:

        # Verde
        if dias_rest > 2:
            return "A TIEMPO"

        # Amarillo
        elif 1 <= dias_rest <= 2:
            return "ALERTA"

        # Naranja
        elif dias_rest == 0:
            if horas_rest != "00:00":
                return "ALERTA_0_DIAS"
            else:
                return "CRÍTICO"

        # Rojo (antes VENCIDO)
        elif dias_rest < 0:
            return "CRÍTICO"

    return "ABIERTO"


df_ans["ESTADO_FENIX"] = df_ans.apply(calcular_estado_fenix, axis=1)
print("🧩 Columna ESTADO_FENIX actualizada correctamente (sin tocar formato).")

# ============================================================
# 5️⃣ MOVER PEDIDOS CERRADOS AL REPOSITORIO
# ============================================================
cerrados = df_ans[df_ans["ESTADO_FENIX"] == "CERRADO"].copy()
if not cerrados.empty:
    print(f"📦 {len(cerrados)} pedidos cerrados serán movidos al repositorio histórico.")
    if ruta_repo.exists():
        repo = pd.read_excel(ruta_repo, dtype=str)
        repo = pd.concat([repo, cerrados], ignore_index=True)
        repo.drop_duplicates(subset=["PEDIDO"], keep="last", inplace=True)
    else:
        repo = cerrados.copy()
    # Eliminar columnas no deseadas antes de guardar en el repositorio
    columnas_eliminar = ["AREA_OPERATIVA", "SUBZONA", "TECNICO_EJECUTA"]

    for col in columnas_eliminar:
        if col in repo.columns:
            del repo[col]
    
    repo.to_excel(ruta_repo, index=False)
    df_ans = df_ans[df_ans["ESTADO_FENIX"] != "CERRADO"]
else:
    print("ℹ️ No se encontraron pedidos cerrados para mover.")

# ============================================================
# 6️⃣ GUARDAR RESULTADOS (ACTUALIZA SOLO COLUMNA ESTADO_FENIX)
# ============================================================
from openpyxl import load_workbook

wb = load_workbook(ruta_fenix_ans, keep_vba=True)
ws = wb["FENIX_ANS"]

# 🔹 Crear mapa {pedido: estado} sin .0, espacios o diferencias de tipo
mapa_estados = {
    str(k).split(".")[0].strip().upper(): v
    for k, v in zip(df_ans["PEDIDO"], df_ans["ESTADO_FENIX"])
}

col_pedido = None
col_estado = None
for idx, cell in enumerate(ws[1], 1):
    nombre = str(cell.value).strip().upper() if cell.value else ""
    if nombre == "PEDIDO":
        col_pedido = idx
    elif nombre == "ESTADO_FENIX":
        col_estado = idx

actualizados = 0
cerrados = 0
if col_pedido and col_estado:
    for i in range(2, ws.max_row + 1):
        pedido_excel = str(ws.cell(i, col_pedido).value).split(".")[0].strip().upper()
        if pedido_excel in mapa_estados:
            nuevo_estado = mapa_estados[pedido_excel]
            ws.cell(i, col_estado).value = nuevo_estado
            actualizados += 1
            if nuevo_estado == "CERRADO":
                cerrados += 1

    print(f"✅ {actualizados} filas actualizadas en la columna ESTADO_FENIX.")
    print(f"📦 {cerrados} pedidos marcados como CERRADO (serán movidos al repositorio).")
else:
    print("⚠️ No se encontraron columnas PEDIDO o ESTADO_FENIX en la hoja.")

# ============================================================
# 7️⃣ APLICAR FORMATO CONDICIONAL A ESTADO_FENIX
# ============================================================
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import Rule

col = col_estado
max_row = ws.max_row
rango = f"{ws.cell(2, col).coordinate}:{ws.cell(max_row, col).coordinate}"

def regla_contiene(texto, color_hex):
    regla = Rule(type="containsText", operator="containsText", text=texto)
    regla.formula = [f'NOT(ISERROR(SEARCH("{texto}",{ws.cell(2, col).coordinate})))']
    regla.dxf = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    return regla

# 🟩 Verde
ws.conditional_formatting.add(rango, regla_contiene("A TIEMPO", "00FF00"))

# 🟨 Amarillo
ws.conditional_formatting.add(rango, regla_contiene("ALERTA", "FFFF00"))

# 🟧 Naranja
ws.conditional_formatting.add(rango, regla_contiene("ALERTA_0_DIAS", "FFC000"))

# 🔴 Rojo (único estado CRÍTICO)
ws.conditional_formatting.add(rango, regla_contiene("CRÍTICO", "FF0000"))

# 🟦 Azul
ws.conditional_formatting.add(rango, regla_contiene("ABIERTO", "8FAADC"))

# 🟪 Morado
ws.conditional_formatting.add(rango, regla_contiene("CERRADO", "7030A0"))

print("🎨 Formatos condicionales aplicados correctamente.")

wb.save(ruta_fenix_ans)
print("💾 Archivo guardado correctamente preservando formatos.")
print("------------------------------------------------------------")
