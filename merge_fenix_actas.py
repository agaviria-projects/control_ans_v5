"""
------------------------------------------------------------
ðŸ”„ MERGE_FENIX_ACTAS.PY â€“ Cruce ProgramaciÃ³n vs Actas (VersiÃ³n Final)
------------------------------------------------------------
Autor: HÃ©ctor A. Gaviria + IA (2025)
------------------------------------------------------------
DescripciÃ³n:
1ï¸âƒ£ Cruza ProgramaciÃ³n (pendientes) vs Actas de Clientes.
2ï¸âƒ£ Actualiza columna ESTADO_FENIX directamente en FENIX_ANS.xlsx.
3ï¸âƒ£ Mueve pedidos cerrados (Ejecutado en Campo + Cumplido)
    al archivo REPOSITORIO_PEDIDOS_CERRADOS.xlsx.
4ï¸âƒ£ Aplica formato de color en ESTADO_FENIX segÃºn dÃ­as restantes.
------------------------------------------------------------
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill


# ------------------------------------------------------------
# ðŸ“‚ RUTAS DE ARCHIVOS
# ------------------------------------------------------------
base_dir = Path(__file__).resolve().parent
ruta_programacion = list(base_dir.glob("data_raw/*pendientes*.*"))
ruta_actas = list(base_dir.glob("data_raw/*Acta_Clientes*.*"))
ruta_fenix_ans = base_dir / "data_clean" / "FENIX_ANS.xlsx"
ruta_repo = base_dir / "data_clean" / "REPOSITORIO_PEDIDOS_CERRADOS.xlsx"

print("------------------------------------------------------------")
print("ðŸ”„ INICIANDO CRUCE PROGRAMACIÃ“N VS ACTAS")
print("------------------------------------------------------------")

if not ruta_programacion or not ruta_actas:
    print("âš ï¸ No se encontraron archivos pendientes o actas en data_raw.")
    exit(1)

archivo_prog = max(ruta_programacion, key=lambda f: f.stat().st_mtime)
archivo_actas = max(ruta_actas, key=lambda f: f.stat().st_mtime)
print(f"ðŸ“˜ ProgramaciÃ³n: {archivo_prog.name}")
print(f"ðŸ“— Actas: {archivo_actas.name}")

# ------------------------------------------------------------
# ðŸ§® LECTOR UNIVERSAL
# ------------------------------------------------------------
def leer_archivo(ruta):
    ext = ruta.suffix.lower()
    if ext in [".csv", ".txt"]:
        try:
            with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
                primera = f.readline()
            if "|" in primera:
                sep = "|"
            elif ";" in primera:
                sep = ";"
            else:
                sep = ","
            df = pd.read_csv(ruta, sep=sep, dtype=str, encoding="utf-8", on_bad_lines="skip")
        except Exception:
            df = pd.read_csv(ruta, sep=sep, dtype=str, encoding="latin1", on_bad_lines="skip")
    elif ext in [".xlsx", ".xls"]:
        df = pd.read_excel(ruta, dtype=str)
    else:
        raise ValueError(f"âŒ Tipo de archivo no soportado: {ruta.name}")
    return df

# ------------------------------------------------------------
# ðŸ§© CARGAR ARCHIVOS
# ------------------------------------------------------------
df_prog = leer_archivo(archivo_prog)
df_actas = leer_archivo(archivo_actas)
df_fenix = pd.read_excel(ruta_fenix_ans, sheet_name="FENIX_ANS", dtype=str)

for df in [df_prog, df_actas, df_fenix]:
    df.columns = df.columns.str.strip().str.lower()

# ------------------------------------------------------------
# ðŸ§© CRUCE DE PEDIDOS
# ------------------------------------------------------------
pedidos_cumplidos = set(df_actas["pedido"].dropna().unique())
df_prog["estado_cruce"] = df_prog["pedido"].apply(
    lambda x: "CUMPLIDO" if x in pedidos_cumplidos else "PENDIENTE"
)

# ------------------------------------------------------------
# ðŸ”— ACTUALIZAR FENIX_ANS (sin perder formato ni estilos)
# ------------------------------------------------------------
if "pedido" in df_fenix.columns:
    print("ðŸ“— Actualizando columna ESTADO_FENIX preservando formato...")

    mapa_estados = dict(zip(df_prog["pedido"], df_prog["estado_cruce"]))

    wb = load_workbook(ruta_fenix_ans)
    ws = wb["FENIX_ANS"]

    columna_estado = None
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(1, col).value).strip().upper() == "ESTADO_FENIX":
            columna_estado = col
            break

    if columna_estado:
        actualizados = 0
        for i in range(2, ws.max_row + 1):
            pedido_excel = str(ws.cell(i, 1).value).strip()  # Columna 1 = pedido
            if pedido_excel in mapa_estados:
                ws.cell(i, columna_estado).value = mapa_estados[pedido_excel]
                actualizados += 1
        print(f"ðŸ’¾ {actualizados} filas actualizadas correctamente en ESTADO_FENIX.")
    else:
        print("âš ï¸ No se encontrÃ³ columna ESTADO_FENIX en la hoja FENIX_ANS.")

    wb.save(ruta_fenix_ans)
    print("âœ… Archivo actualizado preservando estilos, colores y formato condicional.\n")
else:
    print("âš ï¸ No se encontrÃ³ columna 'pedido' en FENIX_ANS.xlsx.")
    exit(1)
# ------------------------------------------------------------
# ðŸ“¦ MOVER PEDIDOS CERRADOS AL REPOSITORIO (flujo limpio)
# ------------------------------------------------------------
print("ðŸ” Verificando coincidencias antes de mover al repositorio...")

df_fenix_actualizado = pd.read_excel(ruta_fenix_ans, sheet_name="FENIX_ANS", dtype=str)
df_fenix_actualizado.columns = df_fenix_actualizado.columns.str.strip().str.lower()

# ðŸ”Ž Buscar pedidos ejecutados en campo y cumplidos
cerrados = df_fenix_actualizado[
    (df_fenix_actualizado["reporte_tecnico"].str.upper().str.contains("EJECUTADO", na=False))
    & (df_fenix_actualizado["estado_fenix"].str.upper() == "CUMPLIDO")
].copy()

if not cerrados.empty:
    print(f"ðŸ“¦ {len(cerrados)} pedidos cerrados serÃ¡n movidos al repositorio...")

    # Normalizar columnas
    cerrados.columns = cerrados.columns.str.strip().str.lower()
    cerrados = cerrados.loc[:, ~cerrados.columns.duplicated()]

    # ðŸ“ Si el repositorio existe, combinar datos
    # ðŸ“ Si el repositorio existe, combinar datos
    if ruta_repo.exists():
        repo = pd.read_excel(ruta_repo, dtype=str)
        repo.columns = repo.columns.str.strip().str.lower()
        repo = repo.loc[:, ~repo.columns.duplicated()]

        # ðŸ§¹ LIMPIEZA ESPECIAL: eliminar filas vacÃ­as y encabezados viejos
        repo = repo[repo.notna().any(axis=1)]
        repo = repo[~repo.apply(lambda fila: any(str(x).strip().lower() in repo.columns for x in fila.values), axis=1)]
        repo.reset_index(drop=True, inplace=True)

        # ðŸ”¹ Combinar datos sin desordenar columnas y sin filas vacÃ­as
        columnas_repo = list(repo.columns)
        columnas_nuevas = [col for col in cerrados.columns if col not in columnas_repo]

        repo = repo.reindex(columns=columnas_repo + columnas_nuevas)
        cerrados = cerrados.reindex(columns=repo.columns)

       # ðŸ§¹ LIMPIEZA ROBUSTA PARA EVITAR FILA VACÃA
        repo.replace(["nan", "None", None], "", inplace=True)

        # Elimina filas que sean 100% vacÃ­as o solo espacios
        repo = repo[repo.apply(lambda fila: ''.join(str(v).strip() for v in fila.values) != '', axis=1)]

        repo.reset_index(drop=True, inplace=True)


    else:
        repo = cerrados.copy()

    # ðŸ§¹ Limpieza final del repositorio
    repo.drop_duplicates(subset=["pedido"], keep="last", inplace=True)
    repo.dropna(axis=1, how="all", inplace=True)
    # ðŸ’¾ GUARDAR REPOSITORIO SIN FILA VACÃA
       

    # ðŸ’¾ GUARDAR REPOSITORIO SIN FILA VACÃA

    # 1ï¸âƒ£ Crear archivo completamente vacÃ­o SIN hoja inicial
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)  # â† EL PASO CLAVE: eliminar la hoja vacÃ­a que crea openpyxl
    wb.create_sheet("REPOSITORIO_CERRADOS")
    wb.save(ruta_repo)

    # 2ï¸âƒ£ Concatenar cerrados ANTES de limpiar
    repo = pd.concat([repo, cerrados], ignore_index=True)

    # 3ï¸âƒ£ Limpieza robusta (eliminar filas realmente vacÃ­as)
    repo.replace(["nan", "None", None], "", inplace=True)
    repo = repo[repo.apply(lambda fila: ''.join(str(v).strip() for v in fila.values) != '', axis=1)]
    repo.drop_duplicates(subset=["pedido"], keep="last", inplace=True)
    repo.reset_index(drop=True, inplace=True)

    # 4ï¸âƒ£ Guardar de forma limpia sin dejar filas fantasma
    with pd.ExcelWriter(ruta_repo, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        repo.to_excel(writer, sheet_name="REPOSITORIO_CERRADOS", index=False)

    print(f"ðŸ’¾ Archivo actualizado: {ruta_repo}")
  
    # ðŸ§¹ Eliminar pedidos cerrados del archivo FENIX_ANS
    print("ðŸ§¹ Eliminando filas cerradas directamente en FENIX_ANS...")
    wb = load_workbook(ruta_fenix_ans)
    ws = wb["FENIX_ANS"]

    # Detectar columna PEDIDO
    col_pedido = None
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(1, col).value).strip().lower() == "pedido":
            col_pedido = col
            break

    pedidos_cerrados = set(cerrados["pedido"].dropna().astype(str))
    filas_eliminadas = 0

    # Eliminar de abajo hacia arriba
    for i in range(ws.max_row, 1, -1):
        pedido_excel = str(ws.cell(i, col_pedido).value).strip()
        if pedido_excel in pedidos_cerrados:
            ws.delete_rows(i, 1)
            filas_eliminadas += 1

    wb.save(ruta_fenix_ans)
    print(f"âœ… {filas_eliminadas} filas eliminadas correctamente de FENIX_ANS.")
    print("------------------------------------------------------------")

    # ðŸ”„ Recargar archivo actualizado antes del formato condicional
    df_fenix_actualizado = pd.read_excel(ruta_fenix_ans, sheet_name="FENIX_ANS", dtype=str)
    df_fenix_actualizado.columns = df_fenix_actualizado.columns.str.strip().str.lower()

else:
    print("â„¹ï¸ No hay pedidos cerrados nuevos para mover al repositorio.")

# ------------------------------------------------------------
# ðŸŽ¨ FORMATO CONDICIONAL Y LÃ“GICA DE ESTADOS
# ------------------------------------------------------------
print("ðŸŽ¨ Aplicando formato condicional en FENIX_ANS...")

wb = load_workbook(ruta_fenix_ans)
ws = wb["FENIX_ANS"]

cols = {str(cell.value).strip().upper(): idx + 1 for idx, cell in enumerate(ws[1])}
col_dias = cols.get("DIAS_RESTANTES")
col_reporte = cols.get("REPORTE_TECNICO")
col_estado = cols.get("ESTADO_FENIX")

if not all([col_dias, col_reporte, col_estado]):
    print("âš ï¸ No se encontraron todas las columnas necesarias para aplicar formato condicional.")
    print(f"   col_dias={col_dias}, col_reporte={col_reporte}, col_estado={col_estado}")
else:
    print(f"ðŸŽ¨ Columnas detectadas correctamente â†’ REPORTE: {col_reporte}, ESTADO: {col_estado}")

# ðŸŽ¨ Colores
verde = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
naranja = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
gris = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# ðŸ”„ Aplicar reglas de negocio
for fila in range(2, ws.max_row + 1):
    try:
        # âœ… No tocar los pedidos que ya estÃ¡n cumplidos
        if str(ws.cell(fila, col_estado).value).strip().upper() == "CUMPLIDO":
            continue

        reporte = str(ws.cell(fila, col_reporte).value).strip().upper()
        dias_texto = str(ws.cell(fila, col_dias).value)
        celda_estado = ws.cell(fila, col_estado)

        # 1ï¸âƒ£ Si el tÃ©cnico no ha reportado nada:
        if reporte == "SIN DATO" or reporte == "":
            celda_estado.value = "ABIERTO"
            celda_estado.fill = gris
            continue

        # 2ï¸âƒ£ Si ya estÃ¡ ejecutado en campo:
        if "EJECUTADO" in reporte:
            dias_num = 0
            if "dÃ­a" in dias_texto:
                try:
                    dias_num = int(dias_texto.split("dÃ­a")[0].strip())
                except:
                    dias_num = 0

            if dias_num > 2:
                celda_estado.value = "A TIEMPO"
                celda_estado.fill = verde
            elif 0 < dias_num <= 2:
                celda_estado.value = "ALERTA"
                celda_estado.fill = amarillo
            elif dias_num == 0 and "hora" in dias_texto:
                celda_estado.value = "A CERO"
                celda_estado.fill = naranja
            elif dias_num < 0:
                celda_estado.value = "VENCIDO"
                celda_estado.fill = rojo
            else:
                celda_estado.value = "ALERTA"
                celda_estado.fill = amarillo
        else:
            celda_estado.value = "ABIERTO"
            celda_estado.fill = gris

    except Exception as e:
        print(f"âš ï¸ Error procesando fila {fila}: {e}")

wb.save(ruta_fenix_ans)
print("âœ… Formato condicional aplicado correctamente.")
print("------------------------------------------------------------")
print("âœ… Cruce, actualizaciÃ³n y formatos finalizados.")
print("------------------------------------------------------------")
